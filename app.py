import os
import sys
import threading
import webbrowser
import json
import time
import fitz
from PIL import Image
from rapidocr_onnxruntime import RapidOCR
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from flask import Flask, render_template, request, url_for, jsonify

# ⬇️ 新增：邮件发送与网络请求必需的依赖库 ⬇️
import smtplib
from email.mime.text import MIMEText
from email.header import Header
from email.utils import formataddr
import urllib.request

# --- 动态获取模板文件夹路径 ---
if getattr(sys, 'frozen', False):
    template_folder = os.path.join(sys._MEIPASS, 'templates')
else:
    template_folder = 'templates'

app = Flask(__name__, template_folder=template_folder)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['STATIC_FOLDER'] = 'static'

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['STATIC_FOLDER'], exist_ok=True)

# 初始化 OCR 引擎
ocr_engine = RapidOCR()

@app.route('/', methods=['GET', 'POST'])
def index():
    image_url = None
    ocr_data_json = "[]"
    orig_w, orig_h = 0, 0
    safe_base_name = ""  # 默认名字
    
    # ==============================
    # 🌟 处理从“历史下拉菜单”加载图纸的请求 (仅限 GET)
    # ==============================
    if request.method == 'GET':
        load_file = request.args.get('load')
        if load_file:
            safe_base_name = "".join([c for c in load_file if c not in r'\/:*?"<>|'])
            image_filename = f"{safe_base_name}.png"
            json_filename = f"{safe_base_name}.json"
            image_path = os.path.join(app.config['STATIC_FOLDER'], image_filename)
            json_path = os.path.join(app.config['STATIC_FOLDER'], json_filename)
            
            # 只要后端有缓存，直接秒开
            if os.path.exists(image_path) and os.path.exists(json_path):
                print(f"⚡ 历史记录加载！正在极速打开 [{safe_base_name}]")
                with open(json_path, 'r', encoding='utf-8') as f:
                    ocr_data_json = f.read()
                img = Image.open(image_path)
                orig_w, orig_h = img.size
                image_url = url_for('static', filename=image_filename)
                
                return render_template('index.html', 
                                       image_url=image_url, ocr_data=ocr_data_json,
                                       orig_w=orig_w, orig_h=orig_h,
                                       filename=safe_base_name)
    
    # ==============================
    # 正常的 POST 上传与初次识别逻辑
    # ==============================
    if request.method == 'POST':
        if 'pdf_file' not in request.files:
            return "没有选择文件"
        file = request.files['pdf_file']
        
        if file.filename != '':
            base_name = os.path.splitext(file.filename)[0]
            safe_base_name = "".join([c for c in base_name if c not in r'\/:*?"<>|'])
            
            image_filename = f"{safe_base_name}.png"
            json_filename = f"{safe_base_name}.json"
            image_path = os.path.join(app.config['STATIC_FOLDER'], image_filename)
            json_path = os.path.join(app.config['STATIC_FOLDER'], json_filename)
            
            # 命中缓存
            if os.path.exists(image_path) and os.path.exists(json_path):
                print(f"⚡ 极速缓存！检测到曾识别过 [{file.filename}]")
                with open(json_path, 'r', encoding='utf-8') as f:
                    ocr_data_json = f.read()
                img = Image.open(image_path)
                orig_w, orig_h = img.size
                image_url = url_for('static', filename=image_filename)
                
                return render_template('index.html', 
                                       image_url=image_url, ocr_data=ocr_data_json,
                                       orig_w=orig_w, orig_h=orig_h,
                                       filename=safe_base_name)
            
            # 首次识别
            pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(pdf_path)
            
            pdf_doc = fitz.open(pdf_path)
            page = pdf_doc.load_page(0) 
            zoom = fitz.Matrix(4, 4)
            pix = page.get_pixmap(matrix=zoom)
            pix.save(image_path)
            image_url = url_for('static', filename=image_filename)
            
            img = Image.open(image_path)
            orig_w, orig_h = img.size
            
            start_t = time.time()
            ocr_result, _ = ocr_engine(image_path)
            
            ocr_data = []
            if ocr_result:
                for item in ocr_result:
                    box = item[0]
                    text = item[1]
                    xs = [p[0] for p in box]
                    ys = [p[1] for p in box]
                    x, y = min(xs), min(ys)
                    w, h = max(xs) - x, max(ys) - y
                    ocr_data.append({'x': float(x), 'y': float(y), 'w': float(w), 'h': float(h), 'text': str(text)})
            
            ocr_data_json = json.dumps(ocr_data, ensure_ascii=False)
            with open(json_path, 'w', encoding='utf-8') as f:
                f.write(ocr_data_json)
                
            return render_template('index.html', 
                                   image_url=image_url, ocr_data=ocr_data_json,
                                   orig_w=orig_w, orig_h=orig_h,
                                   filename=safe_base_name)
            
    # GET 请求（刷新空网页时）
    return render_template('index.html', image_url=image_url, ocr_data=ocr_data_json, orig_w=orig_w, orig_h=orig_h, filename="")

# ==============================
# 📊 高级全息导出 Excel 引擎 (多工作表支持)
# ==============================
@app.route('/api/export', methods=['POST'])
def api_export():
    try:
        payload = request.json
        sheets_data = payload.get('sheets', [])
        export_path = os.path.join(app.config['STATIC_FOLDER'], '红星号管_工程汇总表.xlsx')

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"]) 

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        thick_bottom = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='medium'))

        for sheet_info in sheets_data:
            sheet_name = sheet_info.get('sheet_name', '未命名图纸').strip()
            sheet_name = "".join([c for c in sheet_name if c not in r'[]:*?/\ '])[:31]
            if not sheet_name:
                sheet_name = "未命名"

            table_data = sheet_info.get('data', [])

            base_name = sheet_name
            counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base_name}_{counter}"
                counter += 1

            ws = wb.create_sheet(title=sheet_name)

            for r_idx, row in enumerate(table_data, 1):
                for c_idx, cell_info in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    val = cell_info.get('val', '')
                    color = cell_info.get('color', '')
                    is_header = cell_info.get('is_header', False)
                    is_group_end = cell_info.get('is_group_end', False)

                    cell.value = val
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    if is_header:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        cell.border = thin_border
                    else:
                        cell.font = Font(bold=True) if cell_info.get('bold') else Font()
                        if color and color != 'transparent' and color.startswith('#'):
                            hex_color = color.replace('#', '')
                            if len(hex_color) == 6:
                                cell.fill = PatternFill(start_color="FF"+hex_color, end_color="FF"+hex_color, fill_type="solid")
                        cell.border = thick_bottom if is_group_end else thin_border

            col_widths = [15, 14, 8, 10, 12, 20, 35, 18]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        if not wb.sheetnames: 
            wb.create_sheet("Empty")

        wb.save(export_path)
        return jsonify({'url': url_for('static', filename='红星号管_工程汇总表.xlsx') + "?t=" + str(int(time.time()))})
    except Exception as e:
        print("❌ 导出发生错误:", e)
        return jsonify({'error': str(e)}), 500

# ==============================
# 📧 提交建议反馈引擎 (邮件直连)
# ==============================
@app.route('/api/feedback', methods=['POST'])
def api_feedback():
    try:
        data = request.json
        content = data.get('content', '')
        if not content.strip():
            return jsonify({'error': '内容不能为空'}), 400
        
        # 配置QQ邮箱发送
        sender = '316786168@qq.com'
        password = 'xbqikodgtwfubhhb' # 授权码
        receiver = '316786168@qq.com'
        
        message = MIMEText(f"收到来自天枢·Matrix系统的新反馈：\n\n{content}", 'plain', 'utf-8')
        
        # ⚠️ 修复 QQ 邮箱 550 错误：严格遵循 RFC5322 标准，将昵称与真实邮箱绑定
        message['From'] = formataddr((Header("天枢·Matrix 指令部", 'utf-8').encode(), sender))
        message['To'] = formataddr((Header("系统管理员", 'utf-8').encode(), receiver))
        message['Subject'] = Header("🚨 天枢系统用户建议/Bug报告", 'utf-8')
        
        server = smtplib.SMTP_SSL("smtp.qq.com", 465)
        server.login(sender, password)
        server.sendmail(sender, [receiver], message.as_string())
        server.quit()
        
        return jsonify({'message': '发送成功'})
    except Exception as e:
        print("❌ 邮件发送失败:", str(e))
        return jsonify({'error': str(e)}), 500

# ==============================
# 🔄 GitHub 热更新引擎
# ==============================
@app.route('/api/update', methods=['POST'])
def api_update():
    try:
        github_url = "https://raw.githubusercontent.com/kgdjcf624/Matrix/main/templates/index.html"
        req = urllib.request.Request(github_url, headers={'User-Agent': 'Mozilla/5.0'})
        
        try:
            with urllib.request.urlopen(req) as response:
                html_content = response.read().decode('utf-8')
        except Exception:
            # 尝试 master 分支
            github_url_master = "https://raw.githubusercontent.com/kgdjcf624/Matrix/master/templates/index.html"
            req = urllib.request.Request(github_url_master, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req) as response:
                html_content = response.read().decode('utf-8')
                
        if not html_content or "<html" not in html_content:
            return jsonify({'error': '拉取到的内容无效'}), 400
            
        template_path = os.path.join(app.template_folder, 'index.html')
        with open(template_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        return jsonify({'message': '更新成功，请刷新页面'})
    except Exception as e:
        print("❌ 更新失败:", str(e))
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    def open_browser():
        time.sleep(0) 
        webbrowser.open("http://127.0.0.1:5055")

    threading.Thread(target=open_browser, daemon=True).start()
    app.run(host='127.0.0.1', port=5055, debug=False)